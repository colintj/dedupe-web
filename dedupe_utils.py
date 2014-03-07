import csv
import re
import os
import json
from dedupe import AsciiDammit
import dedupe
from cStringIO import StringIO
from collections import defaultdict, OrderedDict
import logging
from datetime import datetime
from queue import queuefunc
import pdb
from operator import itemgetter
from csvkit import convert
import xlwt
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'upload_data')

class DedupeFileError(Exception): 
    def __init__(self, message):
        Exception.__init__(self, message)
        self.message = message

class DedupeFileIO(object):
    """ 
    Take an uploaded file, figure out what type it is, convert it to csv
    then save it back as the same format.
    """
    def __init__(self, raw, filename):
        self.file_type = convert.guess_format(raw.filename)
        if self.file_type not in ['xls', 'csv', 'xlsx']:
            raise DedupeFileError('%s is not a supported format' % self.file_type)
        self.converted = convert.convert(raw, self.file_type)
        self.line_count = self.converted.count('\n')
        if self.line_count > 10000:
            raise DedupeFileError('Your file has %s rows and we can only currently handle 10,000.' % self.line_count)
        self.file_path = os.path.abspath(os.path.join(UPLOAD_FOLDER, filename))
        with open(self.file_path, 'wb') as f:
            f.write(self.converted)

    def prepare(self, clustered_dupes):
        self.clustered_dupes = clustered_dupes
        self.cluster_count = self._prepareResults()
        self._prepareUniqueResults()

    def _prepareResults(self):
        """ 
        Prepare deduplicated file for writing to various formats with
        duplicates clustered. 
        """
        cluster_membership = {}
        for cluster_id, cluster in enumerate(self.clustered_dupes):
            for record_id in cluster:
                cluster_membership[record_id] = cluster_id

        unique_record_id = cluster_id + 1
        
        f = open(self.file_path, 'rU')
        reader = csv.reader(f)
 
        heading_row = reader.next()
        heading_row.insert(0, 'Group ID')
    
        rows = []

        for row_id, row in enumerate(reader):
            if row_id in cluster_membership:
                cluster_id = cluster_membership[row_id]
            else:
                cluster_id = unique_record_id
                unique_record_id += 1
            row.insert(0, cluster_id)
            rows.append(row)
        rows = sorted(rows, key=itemgetter(0))
        rows.insert(0, heading_row)
        self.clustered_rows = []
        for row in rows:
            d = OrderedDict()
            for k,v in zip(heading_row, row):
                d[k] = v
            self.clustered_rows.append(d)
        f.close()
        return unique_record_id
 
    def _prepareUniqueResults(self):
        """ """
        cluster_membership = {}
        for (cluster_id, cluster) in enumerate(self.clustered_dupes):
            for record_id in cluster:
                cluster_membership[record_id] = cluster_id
 
        f = open(self.file_path, 'rU')
        reader = csv.reader(f)
 
        rows = [reader.next()]
        seen_clusters = set()
        for row_id, row in enumerate(reader):
            if row_id in cluster_membership: 
                cluster_id = cluster_membership[row_id]
                if cluster_id not in seen_clusters:
                    rows.append(row)
                    seen_clusters.add(cluster_id)
            else:
                rows.append(row)
        self.unique_rows = []
        for row in rows:
            d = OrderedDict()
            for k,v in zip(rows[0], row):
                d[k] = v
            self.unique_rows.append(d)
        f.close()
        return self.unique_rows
    
    def writeCSV(self):
        u_path = '%s-deduped_unique.csv' % self.file_path
        d_path = '%s-deduped.csv' % self.file_path
        unique = open(u_path, 'wb')
        writer = csv.DictWriter(unique, self.unique_rows[0].keys())
        writer.writeheader()
        writer.writerows(self.unique_rows)
        unique.close()
        clusters = open(d_path, 'wb')
        writer = csv.DictWriter(clusters, self.clustered_rows[0].keys())
        writer.writeheader()
        writer.writerows(self.clustered_rows)
        clusters.close()
        return d_path, u_path, self.cluster_count, self.line_count

    def _iterExcel(self, outp_type):
        rows = getattr(self,outp_type)
        header = rows[0].keys()
        for r, row in enumerate(rows):
            for c, key in enumerate(header):
                value = row[key]
                yield r,c,value

    def writeXLS(self):
        u_path = '%s-deduped_unique.xls' % self.file_path
        d_path = '%s-deduped.xls' % self.file_path
        clustered_book = xlwt.Workbook(encoding='utf-8')
        clustered_sheet = clustered_book.add_sheet('Clustered Results')
        for r,c,value in self._iterExcel('clustered_rows'):
            clustered_sheet.write(r,c,label=value)
        clustered_book.save(d_path)
        unique_book = xlwt.Workbook(encoding='utf-8')
        unique_sheet = unique_book.add_sheet('Unique Results')
        for r,c,value in self._iterExcel('unique_rows'):
            unique_sheet.write(r,c,label=value)
        unique_book.save(u_path)
        return d_path, u_path, self.cluster_count, self.line_count
    
    def writeXLSX(self):
        u_path = '%s-deduped_unique.xlsx' % self.file_path
        d_path = '%s-deduped.xlsx' % self.file_path
        d_book = Workbook()
        d_ws = d_book.active
        d_ws.title = 'Clustered Results'
        for r,c,value in self._iterExcel('clustered_rows'):
            col = get_column_letter(c + 1)
            d_ws.cell('%s%s' % (col, r + 1)).value = value
        d_book.save(filename=d_path)
        u_book = Workbook()
        u_ws = u_book.active
        u_ws.title = 'Unique Results'
        for r,c,value in self._iterExcel('unique_rows'):
            col = get_column_letter(c + 1)
            u_ws.cell('%s%s' % (col, r + 1)).value = value
        u_book.save(filename=u_path)
        return d_path, u_path, self.cluster_count, self.line_count

class WebDeduper(object):
    
    def __init__(self, deduper,
            file_io=None, 
            training_data=None,
            recall_weight=2):
        self.file_io = file_io
        self.data_d = self.readData()
        self.deduper = deduper
        self.recall_weight = float(recall_weight)
        self.training_data = training_data
        if training_data:
            self.deduper.readTraining(self.training_data)
            self.deduper.train()
            self.settings_path = '%s-settings.dedupe' % self.file_io.file_path
            self.deduper.writeTraining(self.training_data)
            self.deduper.writeSettings(self.settings_path)

    def dedupe(self):
        threshold = self.deduper.threshold(self.data_d, recall_weight=self.recall_weight)
        clustered_dupes = self.deduper.match(self.data_d, threshold)
        self.file_io.prepare(clustered_dupes)
        if self.file_io.file_type == 'csv':
            deduped, deduped_unique, cluster_count, line_count = self.file_io.writeCSV()
        if self.file_io.file_type == 'xls':
            deduped, deduped_unique, cluster_count, line_count = self.file_io.writeXLS()
        if self.file_io.file_type == 'xlsx':
            deduped, deduped_unique, cluster_count, line_count = self.file_io.writeXLSX()
        files = {
            'deduped': deduped,
            'deduped_unique': deduped_unique,
            'cluster_count': cluster_count, 
            'line_count': line_count,
        }
        if self.training_data:
            files['training'] = os.path.relpath(self.training_data, __file__)
            files['settings'] = os.path.relpath(self.settings_path, __file__)
        logger.info(files)
        return files
    
    def preProcess(self, column):
        column = AsciiDammit.asciiDammit(column)
        column = re.sub('  +', ' ', column)
        column = re.sub('\n', ' ', column)
        column = column.strip().strip('"').strip("'").lower().strip()
        return column
 
    def readData(self):
        data = {}
        f = open(self.file_io.file_path, 'rU')
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            clean_row = [(k, self.preProcess(v)) for (k,v) in row.items()]
            row_id = i
            data[row_id] = dedupe.core.frozendict(clean_row)
        return data
    
@queuefunc
def dedupeit(**kwargs):
    d = dedupe.Dedupe(kwargs['field_defs'], kwargs['data_sample'])
    deduper = WebDeduper(d, 
        file_io=kwargs['file_io'],
        training_data=kwargs['training_data'])
    files = deduper.dedupe()
    d.pool.terminate()
    del d
    return files

@queuefunc
def static_dedupeit(**kwargs):
    d = dedupe.StaticDedupe(kwargs['settings_path'])
    deduper = WebDeduper(d, 
        file_io=kwargs['file_io'],
        recall_weight=kwargs['recall_weight'])
    files = deduper.dedupe()
    d.pool.terminate()
    del d
    return files
